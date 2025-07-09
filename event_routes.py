import os
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file # Added send_file
from flask_login import login_required, current_user
from datetime import datetime, timezone
from calendar import Calendar, monthrange
from sqlalchemy import func, desc # Added desc
from sqlalchemy.orm import joinedload, selectinload # Added selectinload
import io # Added io
from openpyxl import Workbook # Added Workbook
from openpyxl.styles import Font, Alignment, PatternFill # Added for Excel styling
import qrcode # Added qrcode
import qrcode.image.svg # Added for SVG QR codes

from extensions import db
from models import Event, User, Rating, Category, Registration, Notification, Role, RegistrationStatus # Corrected import
from forms import EventForm, RatingForm
from utils import save_event_poster

event_bp = Blueprint('main', __name__) # Assuming your event routes are under the 'main' blueprint

@event_bp.route('/')
@event_bp.route('/dashboard')
def dashboard():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))

    page = request.args.get('page', 1, type=int)
    category_id = request.args.get('category', type=int)
    search_query = request.args.get('search', '').strip()
    
    events_query = Event.query.options(joinedload(Event.category)).order_by(Event.start_time.asc())

    if category_id:
        events_query = events_query.filter_by(category_id=category_id)
    if search_query:
        events_query = events_query.filter(Event.title.ilike(f'%{search_query}%'))

    pagination_object = events_query.paginate(page=page, per_page=9, error_out=False)
    categories = Category.query.order_by(Category.name).all()

    return render_template('events/dashboard.html', 
                           events=pagination_object.items, 
                           pagination=pagination_object, 
                           categories=categories,
                           selected_category_id=category_id,
                           search_query=search_query)


@event_bp.route('/event/<int:event_id>')
def view_event(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        flash('Event not found!', 'danger')
        return redirect(url_for('main.dashboard'))
    
    user_registered = False
    user_rating = None
    if current_user.is_authenticated:
        registration = Registration.query.filter_by(user_id=current_user.id, event_id=event.id).first()
        if registration:
            user_registered = True
        user_rating = Rating.query.filter_by(user_id=current_user.id, event_id=event.id).first()

    rating_form = RatingForm()
    
    # Debug prints (keep if still debugging, remove for cleaner output)
    # print(f"DEBUG: Event fetched: {event.title}")
    # try:
    #     print(f"DEBUG: Event ratings relationship type: {type(event.ratings)}")
    #     print(f"DEBUG: Number of ratings for event: {event.ratings.count()}")
    #     print(f"DEBUG: Event average rating: {event.average_rating}")
    # except Exception as e:
    #     print(f"DEBUG: Could not access event.average_rating: {e}")

    return render_template('events/event_detail.html', 
                           event=event, 
                           user_registered=user_registered,
                           user_rating=user_rating,
                           rating_form=rating_form)


@event_bp.route('/event/create', methods=['GET', 'POST'])
@login_required
def create_event():
    if current_user.role not in [Role.ORGANIZER, Role.ADMIN]:
        flash('You do not have permission to create events.', 'danger')
        return redirect(url_for('main.dashboard'))

    form = EventForm()
    # Category choices are populated in EventForm's __init__

    if form.validate_on_submit():
        poster_filename = "default_poster.jpg"
        if form.poster.data:
            poster_filename = save_event_poster(form.poster.data) # Make sure save_event_poster handles default_poster logic
        
        event = Event(
            title=form.title.data,
            description=form.description.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,
            location=form.location.data,
            max_attendees=form.max_attendees.data if form.max_attendees.data is not None else None,
            poster=poster_filename,
            organizer_id=current_user.id,
            category_id=form.category.data
        )
        db.session.add(event)
        db.session.commit()
        flash('Event created successfully!', 'success')
        return redirect(url_for('main.view_event', event_id=event.id))
        
    return render_template('events/create_event.html', form=form, title="Create Event")


@event_bp.route('/event/<int:event_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    event = db.session.get(Event, event_id)
    if not event or (event.organizer_id != current_user.id and current_user.role != Role.ADMIN):
        flash('Event not found or you do not have permission to edit it.', 'danger')
        return redirect(url_for('main.dashboard'))

    form = EventForm(obj=event)
    # Ensure category choices are fresh when editing
    form.category.choices = [(c.id, c.name) for c in Category.query.order_by('name').all()]


    if form.validate_on_submit():
        if form.poster.data:
            event.poster = save_event_poster(form.poster.data)
        
        event.title = form.title.data
        event.description = form.description.data
        event.start_time = form.start_time.data
        event.end_time = form.end_time.data
        event.location = form.location.data
        event.max_attendees = form.max_attendees.data if form.max_attendees.data is not None else None
        event.category_id = form.category.data
        db.session.commit()
        flash('Event updated successfully!', 'success')
        return redirect(url_for('main.view_event', event_id=event.id))
    
    elif request.method == 'GET':
        # These lines ensure form fields are pre-populated on GET request
        form.title.data = event.title
        form.description.data = event.description
        form.start_time.data = event.start_time
        form.end_time.data = event.end_time
        form.location.data = event.location
        form.max_attendees.data = event.max_attendees
        form.category.data = event.category_id
            
    return render_template('events/edit_event.html', form=form, title="Edit Event", event=event)


@event_bp.route('/event/<int:event_id>/delete', methods=['POST'])
@login_required
def delete_event(event_id):
    event = db.session.get(Event, event_id)
    if not event or (event.organizer_id != current_user.id and current_user.role != Role.ADMIN):
        flash('Event not found or you do not have permission to delete it.', 'danger')
        return redirect(url_for('main.dashboard'))

    db.session.delete(event)
    db.session.commit()
    flash('Event deleted successfully!', 'success')
    return redirect(url_for('main.dashboard'))


@event_bp.route('/register_for_event/<int:event_id>', methods=['POST'])
@login_required
def register_for_event(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        flash('Event not found!', 'danger')
        return redirect(url_for('main.dashboard'))

    # FIX: Make event.start_time timezone-aware for comparison
    event_start_time_aware = event.start_time.replace(tzinfo=timezone.utc)
    current_time_aware = datetime.now(timezone.utc)

    if event_start_time_aware < current_time_aware:
        flash('Cannot register for a past event.', 'warning')
        return redirect(url_for('main.view_event', event_id=event.id))

    existing_registration = Registration.query.filter_by(user_id=current_user.id, event_id=event.id).first()
    if existing_registration:
        flash('You are already registered for this event.', 'warning')
        return redirect(url_for('main.view_event', event_id=event.id))

    registration = Registration(user_id=current_user.id, event_id=event.id, status=RegistrationStatus.PENDING)
    db.session.add(registration)
    db.session.commit()

    flash(f'Successfully registered for {event.title}!', 'success')
    return redirect(url_for('main.view_event', event_id=event.id))


@event_bp.route('/unregister_from_event/<int:event_id>', methods=['POST'])
@login_required
def unregister_from_event(event_id):
    registration = Registration.query.filter_by(user_id=current_user.id, event_id=event_id).first()
    if not registration:
        flash('You are not registered for this event.', 'warning')
        return redirect(url_for('main.view_event', event_id=event_id))

    # Optional: Prevent unregistering from past events if desired
    # if registration.event.end_time.replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
    #     flash('Cannot unregister from a past event.', 'warning')
    #     return redirect(url_for('main.view_event', event_id=event_id))

    db.session.delete(registration)
    db.session.commit()
    flash('Successfully unregistered from the event.', 'info')
    return redirect(url_for('main.view_event', event_id=event_id))


@event_bp.route('/event/<int:event_id>/rate', methods=['POST'])
@login_required
def rate_event(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        return redirect(url_for('main.dashboard'))

    # FIX: Make event.end_time timezone-aware before comparison
    event_end_time_aware = event.end_time.replace(tzinfo=timezone.utc)
    current_time_aware = datetime.now(timezone.utc)

    if event_end_time_aware > current_time_aware: # Event must have concluded to be rated
        flash('You can only rate events after they have occurred.', 'warning')
        return redirect(url_for('main.view_event', event_id=event.id))

    form = RatingForm()
    if form.validate_on_submit():
        existing_rating = Rating.query.filter_by(user_id=current_user.id, event_id=event.id).first()
        if existing_rating:
            existing_rating.rating = form.rating.data
            existing_rating.comment = form.comment.data
            flash('Your rating has been updated!', 'success')
        else:
            rating = Rating(
                user_id=current_user.id,
                event_id=event.id,
                rating=form.rating.data,
                comment=form.comment.data
            )
            db.session.add(rating)
            flash('Thank you for rating the event!', 'success')
        db.session.commit()
    return redirect(url_for('main.view_event', event_id=event.id))


@event_bp.route('/event_calendar')
def event_calendar():
    # Get current year and month from query parameters, or default to current date
    year = request.args.get('year', datetime.now(timezone.utc).year, type=int)
    month = request.args.get('month', datetime.now(timezone.utc).month, type=int)

    # Validate year and month to prevent invalid dates
    if not (1 <= month <= 12 and year >= 2000): # Basic validation
        flash('Invalid month or year provided.', 'danger')
        year = datetime.now(timezone.utc).year
        month = datetime.now(timezone.utc).month

    # Get events for the selected month
    start_of_month = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    _, num_days = monthrange(year, month)
    end_of_month = datetime(year, month, num_days, 23, 59, 59, tzinfo=timezone.utc)

    # Fetch events whose start_time falls within the current month, ordered by start time
    events_in_month = Event.query.filter(
        Event.start_time >= start_of_month,
        Event.start_time <= end_of_month
    ).order_by(Event.start_time.asc()).all()

    # Organize events by day for easier template rendering
    events_by_day = {}
    for event in events_in_month:
        day = event.start_time.day
        if day not in events_by_day:
            events_by_day[day] = []
        events_by_day[day].append(event)

    # Generate calendar days (list of lists representing weeks)
    cal = Calendar()
    month_calendar = cal.monthdayscalendar(year, month)

    # For navigation: calculate prev/next month/year
    prev_month_year = (month - 1, year)
    if prev_month_year[0] == 0:
        prev_month_year = (12, year - 1)
    
    next_month_year = (month + 1, year)
    if next_month_year[0] == 13:
        next_month_year = (1, year + 1)
        
    # Month names for display
    month_name = datetime(year, month, 1).strftime('%B')

    return render_template(
        'events/event_calendar.html',
        title=f"Events in {month_name} {year}",
        year=year,
        month=month,
        month_name=month_name,
        month_calendar=month_calendar, # The grid of days
        events_by_day=events_by_day, # Events organized by day
        prev_month=prev_month_year[0],
        prev_year=prev_month_year[1],
        next_month=next_month_year[0],
        next_year=next_month_year[1]
    )


@event_bp.route('/my_registrations')
@login_required
def my_registrations():
    registrations = Registration.query.filter_by(user_id=current_user.id).order_by(Registration.registration_date.desc()).all()
    return render_template('events/my_registrations.html', registrations=registrations)


@event_bp.route('/statistics')
def statistics():
    total_events = db.session.query(func.count(Event.id)).scalar()
    total_users = db.session.query(func.count(User.id)).scalar()
    total_registrations = db.session.query(func.count(Registration.id)).scalar()

    # Calculate overall average rating
    overall_avg_rating_result = db.session.query(func.avg(Rating.rating)).scalar()
    overall_avg_rating = overall_avg_rating_result if overall_avg_rating_result is not None else 0.0
    

    # REMOVE: Data queries for individual charts (category_labels, most_registered_labels, etc.)
    # These are no longer needed if we're not drawing charts.

    return render_template(
        'events/statistics.html',
        title="Application Statistics",
        total_events=total_events,
        total_users=total_users,
        total_registrations=total_registrations,
        overall_avg_rating=overall_avg_rating # Pass overall average
    )

@event_bp.route('/organizer_dashboard')
@login_required
def organizer_dashboard():
    if current_user.role not in [Role.ORGANIZER, Role.ADMIN]:
        flash('You do not have permission to view this page.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    events = Event.query.filter_by(organizer_id=current_user.id).all()
    # Calculate stats for organizer dashboard
    total_organized_events = len(events)
    total_regs_for_organized_events = sum(event.registrations.count() for event in events)
    
    # Calculate average rating across all organized events
    all_organized_ratings = [rating.rating for event in events for rating in event.ratings.all()]
    avg_rating_organized = sum(all_organized_ratings) / len(all_organized_ratings) if all_organized_ratings else 0.0

    return render_template('events/organizer_dashboard.html', 
                           organized_events=events,
                           total_organized_events=total_organized_events,
                           total_regs_for_organized_events=total_regs_for_organized_events,
                           avg_rating_organized=avg_rating_organized)


@event_bp.route('/event/<int:registration_id>/qrcode')
@login_required
def get_qrcode(registration_id):
    registration = db.session.get(Registration, registration_id)
    # Ensure registration exists and belongs to current user
    if not registration or registration.user_id != current_user.id:
        flash('QR Code not found or you do not have permission to view it.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Data to embed in QR code (e.g., registration ID, a unique token)
    qr_data = f"registration_id:{registration.id}:user_id:{registration.user.id}:event_id:{registration.event.id}"
    
    # Generate QR code as SVG
    factory = qrcode.image.svg.SvgPathImage
    qr_img = qrcode.make(qr_data, image_factory=factory)
    
    buffer = io.BytesIO()
    qr_img.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, mimetype='image/svg+xml', download_name=f'qr_reg_{registration.id}.svg')


@event_bp.route('/event/<int:event_id>/check_in', methods=['GET', 'POST'])
@login_required
def check_in(event_id):
    event = db.session.get(Event, event_id)
    # Ensure event exists and current user is organizer or admin
    if not event or (event.organizer_id != current_user.id and current_user.role != Role.ADMIN):
        flash('Event not found or you do not have permission to access check-in for this event.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Fetch all registrations for this event to display them
    registrations = Registration.query.filter_by(event_id=event.id).options(
        selectinload(Registration.user)
    ).order_by(Registration.registration_date.asc()).all()

    if request.method == 'POST':
        # This part handles manual check-in by registration ID or QR data
        reg_id_to_check_in = request.form.get('registration_id_input', type=int)
        qr_data_string = request.form.get('qr_data_input')

        registration_to_update = None
        if reg_id_to_check_in:
            registration_to_update = db.session.get(Registration, reg_id_to_check_in)
        elif qr_data_string:
            # Parse QR data: e.g., "registration_id:X:user_id:Y:event_id:Z"
            parts = qr_data_string.split(':')
            if len(parts) == 6 and parts[0] == 'registration_id' and parts[2] == 'user_id' and parts[4] == 'event_id':
                try:
                    scanned_reg_id = int(parts[1])
                    scanned_event_id = int(parts[5])
                    if scanned_event_id == event.id: # Ensure QR is for this event
                        registration_to_update = db.session.get(Registration, scanned_reg_id)
                    else:
                        flash('QR code is for a different event.', 'danger')
                except ValueError:
                    flash('Invalid QR data format.', 'danger')
            else:
                flash('QR data format not recognized.', 'danger')

        if registration_to_update and registration_to_update.event_id == event.id:
            if registration_to_update.status == RegistrationStatus.APPROVED:
                flash(f'{registration_to_update.user.username} is already checked in.', 'info')
            else:
                registration_to_update.status = RegistrationStatus.APPROVED
                db.session.commit()
                flash(f'Successfully checked in {registration_to_update.user.username}!', 'success')
        else:
            flash('Registration not found or not for this event.', 'danger')
        
        return redirect(url_for('main.check_in', event_id=event.id))
    
    return render_template('events/check_in.html', title=f"Check-in for {event.title}", event=event, registrations=registrations)

@event_bp.route('/event/<int:event_id>/export_registrations')
@login_required
def export_registrations(event_id):
    event = db.session.get(Event, event_id)
    if not event or (event.organizer_id != current_user.id and current_user.role != Role.ADMIN):
        flash('Event not found or you do not have permission to export registrations for this event.', 'danger')
        return redirect(url_for('main.dashboard'))

    registrations = Registration.query.filter_by(event_id=event.id).options(joinedload(Registration.user)).order_by(Registration.registration_date.asc()).all()

    if not registrations:
        flash(f'No participants registered for "{event.title}" to export.', 'info')
        return redirect(url_for('main.view_event', event_id=event.id))

    # All of the Excel generation logic needs to be within this block,
    # so 'output' is guaranteed to be defined if 'registrations' is not empty.
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{event.title} Registrations"

        headers = ["Registration ID", "Attendee Username", "Attendee Email", "Registration Date", "Status"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        row_num = 2
        for reg in registrations:
            ws.cell(row=row_num, column=1, value=reg.id)
            ws.cell(row=row_num, column=2, value=reg.user.username)
            ws.cell(row=row_num, column=3, value=reg.user.email)
            # Use localize_datetime if you want the Excel export time to be localized
            # Otherwise, strftime provides consistent string format
            ws.cell(row=row_num, column=4, value=reg.registration_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=5, value=reg.status.value.title())
            row_num += 1
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO() # 'output' is defined here
        wb.save(output)
        output.seek(0)

        safe_event_title = "".join([c for c in event.title if c.isalnum() or c in (' ', '.', '_')]).replace(' ', '_')
        filename = f"{safe_event_title}_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        # Catch any unexpected errors during Excel generation
        flash(f'An error occurred during export: {e}', 'danger')
        print(f"ERROR: Excel export failed for event {event.id}: {e}") # Print to console for debugging
        return redirect(url_for('main.view_event', event_id=event.id))
@event_bp.route('/event/<int:event_id>/gallery')
def event_gallery(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        flash('Event not found!', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # In a real implementation, you'd fetch EventPhoto objects associated with this event
    # For now, we just pass the event
    
    return render_template('events/event_gallery.html', title=f"Gallery for {event.title}", event=event)